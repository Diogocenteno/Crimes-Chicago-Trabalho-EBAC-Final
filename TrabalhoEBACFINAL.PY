# -*- coding: utf-8 -*-
#%%
"""
Trabalho Final EBAC - ANÁLISE CRIMINAL DE CHICAGO COMPLETA - 2001-2025 A 2030
Autor: Diogo Centeno

"""

import pandas as pd
import os
import matplotlib.pyplot as plt
import seaborn as sns
from fpdf import FPDF
from tabulate import tabulate
from sklearn.linear_model import LinearRegression
import numpy as np
from tqdm import tqdm
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import logging

#%%
# Caminhos dos arquivos no seu computador
caminho_ano_tipo = r"C:/Users/diogo/Desktop/Crimes_po_ano_tipo.xlsx"
caminho_bairro = r"C:/Users/diogo/Desktop/Crimes_por_bairro.xlsx"
caminho_local = r"C:/Users/diogo/Desktop/Local_dos_crimes.xlsx"

# Leitura dos arquivos Excel
df_ano_tipo = pd.read_excel(caminho_ano_tipo)
df_bairro = pd.read_excel(caminho_bairro)
df_local = pd.read_excel(caminho_local)

# Visualização das primeiras linhas de cada DataFrame
print("Crimes por ano e tipo:")
print(df_ano_tipo.head(), '\n')

print("Crimes por bairro:")
print(df_bairro.head(), '\n')

print("Local dos crimes:")
print(df_local.head(), '\n')

#%%
# ======================================================
# CONFIGURAÇÃO INICIAL
# ======================================================
# Configuração do logger
PASTA_RESULTADOS = "resultados"  # ou o caminho que você quiser

# Cria a pasta se não existir
os.makedirs(PASTA_RESULTADOS, exist_ok=True)

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler(os.path.join(PASTA_RESULTADOS, 'analise_criminal.log')),
        logging.StreamHandler()
    ]
)

logging.info("="*60)
logging.info("🔍 INÍCIO DA ANÁLISE CRIMINAL")
logging.info("="*60)

sns.set_theme(style="darkgrid", palette="viridis")
COLOR_PRIMARY = '#2ecc71'
COLOR_SECONDARY = '#34495e'

# Caminhos
DESKTOP = os.path.join(os.path.expanduser("~"), "Desktop")
PASTA_RESULTADOS = os.path.join(DESKTOP, "Analise_Criminal_Resultados")
ARQUIVO_FINAL = os.path.join(PASTA_RESULTADOS, "Dados_Consolidados.xlsx")

# Criar pasta resultados se não existir
os.makedirs(PASTA_RESULTADOS, exist_ok=True)

#%%
# ======================================================
# CARREGAMENTO DE DADOS
# ======================================================
print("\n📥 CARREGAMENTO DE DADOS".center(100))

def carregar_dados():
    try:
        logging.info("Iniciando carregamento dos dados...")
        df_ano = pd.read_excel(os.path.join(DESKTOP, "Crimes_po_ano_tipo.xlsx"), sheet_name="Crimes_por_ano_tipo")
        df_bairro = pd.read_excel(os.path.join(DESKTOP, "Crimes_por_bairro.xlsx"), sheet_name="Crimes_por_bairro")
        df_local = pd.read_excel(os.path.join(DESKTOP, "Local_dos_crimes.xlsx"), sheet_name="Localizacao_natureza")

        df_ano["year"] = df_ano["year"].astype(int)
        df_local["year"] = df_local["year"].fillna(0).astype(int)

        for df, nome in zip([df_ano, df_bairro, df_local], ["Anual", "Bairro", "Local"]):
            if df.empty:
                raise ValueError(f"Dataset {nome} está vazio")

        logging.info("✅ Dados carregados com sucesso!")
        return df_ano, df_bairro, df_local
    except Exception as e:
        logging.error(f"❌ Erro no carregamento: {e}")
        exit(1)


df_ano, df_bairro, df_local = carregar_dados()

#%%
# ======================================================
# PRÉ-VISUALIZAÇÃO DOS DADOS
# ======================================================
def mostrar_amostra(df, nome, n=3):
    print(f"\n📑 Amostra de {nome} ({len(df)} registros):")
    print(tabulate(df.head(n), headers='keys', tablefmt='psql', showindex=False))

mostrar_amostra(df_ano, "Dados Anuais")
mostrar_amostra(df_bairro, "Dados por Bairro")
mostrar_amostra(df_local, "Dados de Localização")

#%%
# ======================================================
# PROCESSAMENTO E CRUZAMENTOS
# ======================================================
try:
    merged = pd.merge(
        pd.merge(df_ano, df_bairro, on="primary_type", how="left", suffixes=("_ano", "_bairro")),
        df_local, on=["primary_type", "year"], how="left"
    )
    merged.fillna({
        'ward': 'Não informado',
        'community_area': 'Não informado',
        'location_description': 'Não especificado'
    }, inplace=True)
    mostrar_amostra(merged, "Dados Cruzados")
except Exception as e:
    print(f"\n❌ Erro no cruzamento: {e}")
    exit()

#%%
# ======================================================
# ANÁLISES ESTATÍSTICAS
# ======================================================
def analise_completa():
    try:
        evolucao = merged.groupby('year')['total_crimes'].sum().reset_index()
        crimes = merged.groupby('primary_type')['total_crimes'].sum().nlargest(5)
        locais = merged.groupby('location_description')['total_crimes'].sum().nlargest(5)
        if 'arrest' in merged.columns:
            eficiencia = merged.groupby('ward').agg(
                total_prisoes=('arrest', 'sum'),
                taxa_prisao=('arrest', 'mean')
            ).nlargest(5, 'taxa_prisao')
        else:
            eficiencia = pd.DataFrame()
        return {'evolucao': evolucao, 'crimes': crimes, 'locais': locais, 'eficiencia': eficiencia}
    except Exception as e:
        print(f"\n❌ Erro nas análises: {e}")
        return {'evolucao': pd.DataFrame(), 'crimes': pd.Series(dtype=float), 'locais': pd.Series(dtype=float), 'eficiencia': pd.DataFrame()}

resultados_analise = analise_completa()

#%%
# ======================================================
# VISUALIZAÇÕES AUXILIARES
# ======================================================
def criar_grafico_linha_tendencia(df):
    plt.figure(figsize=(12, 6))
    top_crimes = df.groupby('primary_type')['total_crimes'].sum().nlargest(5).index
    df_filtered = df[df['primary_type'].isin(top_crimes)]
    
    sns.lineplot(
        data=df_filtered,
        x='year',
        y='total_crimes',
        hue='primary_type',
        style='primary_type',
        markers=True,
        dashes=False,
        linewidth=2.5,
        markersize=8
    )
    
    plt.title('Evolução Temporal dos Principais Tipos de Crime', fontsize=14, pad=20)
    plt.xlabel('Ano', fontsize=12)
    plt.ylabel('Total de Ocorrências', fontsize=12)
    plt.legend(title='Tipo de Crime', bbox_to_anchor=(1.05, 1), loc='upper left')
    plt.tight_layout()
    caminho = os.path.join(PASTA_RESULTADOS, "tendencia_top_crimes.png")
    plt.savefig(caminho, dpi=300)
    plt.close()
    print(f"✅ Gráfico de tendência salvo: {caminho}")

def criar_heatmap_distribuicao(df):
    plt.figure(figsize=(10, 8))
    cross_tab = pd.crosstab(df['primary_type'], df['location_description'].fillna('Não especificado'))
    cross_tab = cross_tab.loc[:, cross_tab.sum().nlargest(10).index]
    
    sns.heatmap(
        cross_tab,
        cmap='rocket',
        annot=True,
        fmt='d',
        linewidths=.5,
        annot_kws={"size": 8}
    )
    
    plt.title('Distribuição de Crimes por Local e Tipo', fontsize=14, pad=20)
    plt.xlabel('Localização', fontsize=12)
    plt.ylabel('Tipo de Crime', fontsize=12)
    plt.xticks(rotation=45, ha='right')
    plt.tight_layout()
    caminho = os.path.join(PASTA_RESULTADOS, "heatmap_distribuicao.png")
    plt.savefig(caminho, dpi=300)
    plt.close()
    print(f"✅ Heatmap salvo: {caminho}")

#%%
# ======================================================
# ANÁLISE EXPLORATÓRIA ATUALIZADA
# ======================================================
def analise_exploratoria(df):
    exploratorios = {}

    # Gráfico de distribuição anual atualizado
    plt.figure(figsize=(12, 6))
    ax = sns.barplot(
        data=df.groupby("year")["total_crimes"].sum().reset_index(),
        x="year",
        y="total_crimes",
        color=COLOR_PRIMARY
    )
    
    for p in ax.patches:
        ax.annotate(f"{p.get_height():.0f}", 
                    (p.get_x() + p.get_width() / 2., p.get_height()),
                    ha='center', va='center', 
                    xytext=(0, 9), 
                    textcoords='offset points',
                    fontsize=9)
    
    plt.title("Distribuição Anual de Crimes", fontsize=14, pad=15)
    plt.xlabel("Ano", fontsize=12)
    plt.ylabel("Total de Crimes", fontsize=12)
    plt.xticks(rotation=45)
    plt.tight_layout()
    caminho = os.path.join(PASTA_RESULTADOS, "distribuicao_ano.png")
    plt.savefig(caminho, dpi=300)
    plt.close()
    exploratorios["distribuicao_ano"] = "Distribuição anual de crimes"
    print(f"✅ Gráfico de distribuição anual salvo: {caminho}")

    # Novo gráfico de tendência
    criar_grafico_linha_tendencia(df)
    exploratorios["tendencia_top_crimes"] = "Evolução dos principais tipos de crime"
    
    # Novo heatmap
    criar_heatmap_distribuicao(df)
    exploratorios["heatmap_distribuicao"] = "Relação entre tipo de crime e localização"

    # Boxplot atualizado
    plt.figure(figsize=(10, 6))
    sns.boxplot(
        x=df["total_crimes"],
        color=COLOR_SECONDARY,
        showfliers=False,
        notch=True
    )
    plt.title("Distribuição de Crimes por Registro", fontsize=14, pad=15)
    plt.xlabel("Total de Crimes", fontsize=12)
    plt.tight_layout()
    caminho = os.path.join(PASTA_RESULTADOS, "boxplot.png")
    plt.savefig(caminho, dpi=300)
    plt.close()
    exploratorios["boxplot_crimes"] = "Boxplot da distribuição de crimes"
    print(f"✅ Boxplot salvo: {caminho}")

    # Estatísticas descritivas
    stats = df["total_crimes"].describe()
    print(f"\n📊 Estatísticas descritivas de total_crimes:\n{stats}")

    return exploratorios

resultados_exploratorios = analise_exploratoria(merged)

# ======================================================
# ESTATÍSTICAS DESCRITIVAS
# ======================================================
print("\n📊 Gerando estatísticas descritivas de total_crimes...")
estatisticas_total_crimes = df_ano['total_crimes'].describe().reset_index()
estatisticas_total_crimes.columns = ['Estatística', 'Valor']

#%%
# ======================================================
# PROJEÇÃO ATÉ 2030
# ======================================================
def gerar_projecao_criminalidade(df, salvar_em=None):
    df_ano = df.groupby('year')['total_crimes'].sum().reset_index()
    df_ano = df_ano[df_ano['year'] >= 2001]

    # Modelo
    X = df_ano[['year']]
    y = df_ano['total_crimes']
    modelo = LinearRegression()
    modelo.fit(X, y)

    anos_futuros = np.arange(df_ano['year'].max() + 1, 2031).reshape(-1, 1)
    previsoes = modelo.predict(anos_futuros)

    # DataFrame da projeção
    df_proj = pd.DataFrame({
        "Ano": anos_futuros.flatten(),
        "Projecao_Crimes": previsoes.round().astype(int)
    })

    if salvar_em:
        os.makedirs(salvar_em, exist_ok=True)

        print("📊 Gerando gráfico de projeção...")
        plt.figure(figsize=(12, 6))
        plt.plot(X['year'], y, label="Dados Reais", color='blue', marker='o')
        plt.plot(df_proj["Ano"], df_proj["Projecao_Crimes"], label="Projeção até 2030", linestyle="--", color="orange", marker='x')
        plt.title("Projeção da Criminalidade até 2030", fontsize=14)
        plt.xlabel("Ano")
        plt.ylabel("Total de Crimes")
        plt.legend()
        plt.grid(True)
        plt.tight_layout()

        caminho_grafico = os.path.join(salvar_em, "projecao_crimes.png")
        plt.savefig(caminho_grafico, dpi=300)
        plt.close()
        print(f"✅ Gráfico salvo em: {caminho_grafico}")

        
        # Novo: salvar Excel com formatação
        caminho_excel = os.path.join(salvar_em, "tabela_projecao_2030.xlsx")
        df_proj.to_excel(caminho_excel, index=False)

        wb = load_workbook(caminho_excel)
        ws = wb.active

        header_font = Font(bold=True)
        alignment = Alignment(horizontal="center")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin'),
        )

        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=1, column=col)
            cell.font = header_font
            cell.alignment = alignment
            cell.border = border

        for row in range(2, ws.max_row + 1):
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=row, column=col)
                cell.alignment = alignment
                cell.border = border
                if col == 2:
                    cell.number_format = '#,##0'

        for col in ws.columns:
            max_length = max(len(str(cell.value)) for cell in col if cell.value)
            adjusted_width = max_length + 2
            ws.column_dimensions[get_column_letter(col[0].column)].width = adjusted_width

        wb.save(caminho_excel)
        print(f"✅ Tabela Excel formatada salva em: {caminho_excel}")

    return df_proj
    
# Executar projeção
print("\n🔮 Iniciando projeção da criminalidade...")
caminho_projecao = gerar_projecao_criminalidade(merged, salvar_em=PASTA_RESULTADOS)
resultados_exploratorios["projecao_crimes"] = "Projeção da criminalidade até 2030"

# ======================================================
# ESTATÍSTICAS DESCRITIVAS - TOTAL DE CRIMES (DETALHADA)
# ======================================================
print("\n📊 Gerando estatísticas descritivas detalhadas de total_crimes...")

descricao_estatisticas = {
    'count': ('Contagem', 'Número total de registros de anos analisados', 'Quantidade de pontos de dados disponíveis'),
    'mean': ('Média', 'Média aritmética dos crimes por ano', 'Nível médio de criminalidade anual'),
    'std': ('Desvio Padrão', 'Variação dos crimes entre os anos', 'Quanto os números variam da média'),
    'min': ('Mínimo', 'Ano com menor total de crimes', 'Menor valor registrado no histórico'),
    '25%': ('1º Quartil (25%)', '25% dos anos têm menos que esse total de crimes', 'Limite inferior da média histórica'),
    '50%': ('Mediana (50%)', 'Valor central da distribuição de crimes por ano', '50% dos anos estão abaixo/acima'),
    '75%': ('3º Quartil (75%)', '75% dos anos têm menos que esse total de crimes', 'Limite superior da média histórica'),
    'max': ('Máximo', 'Ano com maior total de crimes', 'Pico histórico de criminalidade anual')
}

estatisticas = df_ano['total_crimes'].describe()
estatisticas_detalhadas = []

for chave, valor in estatisticas.items():
    nome, descricao, interpretacao = descricao_estatisticas.get(chave, (chave, '', ''))
    estatisticas_detalhadas.append({
        'Estatística': nome,
        'Valor': f'{valor:,.2f}',
        'Descrição': descricao,
        'Interpretação Executiva': interpretacao
    })

estatisticas_total_crimes = pd.DataFrame(estatisticas_detalhadas)

#%%
# ======================================================
# SALVAR DADOS CONSOLIDADOS EM EXCEL
# ======================================================
print("\n💾 Salvando dados consolidados em Excel...")
with pd.ExcelWriter(ARQUIVO_FINAL) as writer:
    df_ano.to_excel(writer, sheet_name='Dados_Anuais', index=False)
    df_bairro.to_excel(writer, sheet_name='Dados_Bairro', index=False)
    df_local.to_excel(writer, sheet_name='Dados_Localizacao', index=False)
    merged.to_excel(writer, sheet_name='Dados_Integrados', index=False)
    for nome, dados in resultados_analise.items():
        if not dados.empty:
            if isinstance(dados, pd.Series):
                dados = dados.reset_index()
            dados.to_excel(writer, sheet_name=nome.capitalize(), index=False)
# 🔁 Adiciona a planilha com a projeção
    caminho_projecao.to_excel(writer, sheet_name='Projecao_2030', index=False)
    estatisticas_total_crimes.to_excel(writer, sheet_name='Estatisticas_Total', index=False)
print(f"✅ Arquivo Excel salvo: {ARQUIVO_FINAL}")

#%%
# Caminhos dos arquivos
caminho_ano_tipo = r"C:/Users/diogo/Desktop/Crimes_po_ano_tipo.xlsx"
caminho_bairro = r"C:/Users/diogo/Desktop/Crimes_por_bairro.xlsx"
caminho_local = r"C:/Users/diogo/Desktop/Local_dos_crimes.xlsx"

# Carregar os dados
df_ano_tipo = pd.read_excel(caminho_ano_tipo)
df_bairro = pd.read_excel(caminho_bairro)
df_local = pd.read_excel(caminho_local)

# Limpar nomes das colunas
df_ano_tipo.columns = df_ano_tipo.columns.str.strip()
df_bairro.columns = df_bairro.columns.str.strip()
df_local.columns = df_local.columns.str.strip()

# Mesclar ano_tipo com local por 'primary_type' e 'year'
df_temp = df_ano_tipo.merge(df_local, on=["primary_type", "year"], how="outer")

# Mesclar resultado com bairro por 'primary_type'
df_merged = df_temp.merge(df_bairro, on="primary_type", how="outer")

# Marcar crimes graves
df_merged["CRIME_GRAVE"] = df_merged["primary_type"].apply(
    lambda x: "Sim" if str(x).upper() in ["HOMICÍDIO", "ESTUPRO", "LATROCÍNIO"] else "Não"
)

# Tentar encontrar uma coluna com "total" no nome para calcular a taxa por 1000 hab
coluna_total = next((col for col in df_merged.columns if "total" in col.lower()), None)
if coluna_total:
    try:
        df_merged["CRIMES_POR_1000_HAB"] = df_merged[coluna_total] / (df_merged[coluna_total] / 1000)
    except Exception as e:
        print(f"⚠️ Erro ao calcular CRIMES_POR_1000_HAB com '{coluna_total}': {e}")
else:
    print("⚠️ Nenhuma coluna de total encontrada para calcular CRIMES_POR_1000_HAB.")

# Criar pasta de saída na Área de Trabalho
pasta_saida = os.path.join(os.path.expanduser("~"), "Desktop", "Analise_Criminal_Resultados")
os.makedirs(pasta_saida, exist_ok=True)
caminho_final = os.path.join(pasta_saida, "Crimes_Unificados.xlsx")

# Exportar para Excel
df_merged.to_excel(caminho_final, index=False)
print(f"✅ Arquivo Excel salvo com sucesso em: {caminho_final}")

#%%
# ======================================================
# GERAÇÃO DO RELATÓRIO COMPLETO EM PDF
# ======================================================
class PDFReport(FPDF):
    def chapter_title(self, title):
        self.set_font("Arial", "B", 14)
        self.set_text_color(0, 0, 128)
        self.cell(0, 10, title, 0, 1, "L")
        self.ln(4)

    def chapter_body(self, body_text):
        self.set_font("Arial", "", 12)
        self.set_text_color(0, 0, 0)
        self.multi_cell(0, 8, body_text)
        self.ln(5)

def gerar_relatorio_pdf_completo(resultados_analise, pasta_resultados):

    pdf = PDFReport()
    pdf.set_auto_page_break(auto=True, margin=15)
    pdf.add_page()

    # Título do relatório
    pdf.set_font("Arial", "B", 16)
    pdf.cell(0, 10, "Relatório Completo de Análise Criminal", 0, 1, "C")
    pdf.ln(5)

    # Seção de Introdução
    pdf.chapter_title("Introdução e Importância da Análise")
    introducao = """A análise de dados criminais é uma ferramenta essencial para o desenvolvimento de estratégias eficazes de segurança pública. Através da identificação de padrões temporais, distribuição geográfica e características dos crimes, é possível:

1. Otimizar a alocação de recursos policiais
2. Desenvolver políticas preventivas direcionadas
3. Identificar fatores de risco sociais e ambientais
4. Medir a eficácia de intervenções realizadas
5. Prever tendências futuras para ação proativa

Esta análise combina técnicas estatísticas avançadas com visualização de dados intuitiva para transformar dados brutos em insights acionáveis."""
    pdf.chapter_body(introducao)

    # Seção de Metodologia
    pdf.chapter_title("Metodologia Científica")
    metodologia = """Fluxo de Análise:
1. Coleta de dados de múltiplas fontes
2. Integração e validação dos datasets
3. Análise exploratória inicial
4. Processamento e limpeza dos dados
5. Modelagem estatística básica
6. Visualização de padrões relevantes
7. Interpretação contextualizada dos resultados

Técnicas Utilizadas:
- Análise de séries temporais
- Mapeamento de densidade criminal
- Identificação de outliers estatísticos
- Análise de correlação espacial"""
    pdf.chapter_body(metodologia)

    # Inserir gráficos
    graficos = [
        ("Distribuição Anual de Crimes", os.path.join(pasta_resultados, "distribuicao_ano.png")),
        ("Evolução dos Principais Tipos de Crime", os.path.join(pasta_resultados, "tendencia_top_crimes.png")),
        ("Heatmap de Distribuição de Crimes por Local", os.path.join(pasta_resultados, "heatmap_distribuicao.png")),
        ("Boxplot da Distribuição de Crimes", os.path.join(pasta_resultados, "boxplot.png")),
        ("Projeção de Criminalidade até 2030", os.path.join(pasta_resultados, "projecao_crimes.png"))
    ]

    for titulo, caminho_img in graficos:
        if os.path.exists(caminho_img):
            pdf.chapter_title(titulo)
            pdf.image(caminho_img, x=15, w=180)
            pdf.ln(10)
        else:
            pdf.set_font("Arial", "I", 11)
            pdf.cell(0, 10, f"Gráfico não encontrado: {titulo}", 0, 1)
            pdf.ln(5)

    # Inserir tabelas resumidas
    pdf.chapter_title("Tabelas Resumidas")

    def inserir_tabela(df, titulo):
        if df.empty:
            return
        pdf.set_font("Arial", "B", 12)
        pdf.cell(0, 8, titulo, 0, 1)
        pdf.set_font("Arial", "", 10)
        tabela_str = tabulate(df.head(10), headers='keys', tablefmt='psql', showindex=False)
        for linha in tabela_str.split('\n'):
            pdf.cell(0, 5, linha, ln=1)
        pdf.ln(5)

    if 'crimes' in resultados_analise and not resultados_analise['crimes'].empty:
        top_crimes = resultados_analise['crimes'].reset_index()
        inserir_tabela(top_crimes, "Top 5 Crimes Mais Frequentes")

    if 'locais' in resultados_analise and not resultados_analise['locais'].empty:
        top_locais = resultados_analise['locais'].reset_index()
        inserir_tabela(top_locais, "Top 5 Locais com Mais Crimes")

    if 'eficiencia' in resultados_analise and not resultados_analise['eficiencia'].empty:
        inserir_tabela(resultados_analise['eficiencia'], "Eficiência das Prisões por Região")

    # Fonte dos dados
    pdf.ln(10)
    pdf.set_font("Arial", "I", 10)
    fonte_texto = (
        "Fonte dos dados: Dados oficiais da polícia, compilados e analisados pelo usuário.\n"
        "Link da fonte: https://www.kaggle.com/datasets/utkarshx27/crimes\n"
        "Link da fonte:https://data.cityofchicago.org/browse?sortBy=most_accessed&page=1&pageSize=20\n"
    )
    pdf.multi_cell(0, 8, fonte_texto)

    # Salvar PDF
    arquivo_pdf = os.path.join(pasta_resultados, "Relatorio_Completo.pdf")
    pdf.output(arquivo_pdf)
    print(f"\n✅ Relatório PDF completo salvo em: {arquivo_pdf}")
    
#%%
# Exemplo de uso no seu código principal:
if __name__ == "__main__":
    # ... depois de gerar todos os dados e salvar gráficos ...
    gerar_relatorio_pdf_completo(resultados_analise, PASTA_RESULTADOS)
 
print("\n🏁 Análise finalizada com sucesso.")
print(f"📂 Resultados salvos em: {PASTA_RESULTADOS}")    